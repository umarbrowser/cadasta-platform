import os

from openpyxl import Workbook

from .base import Exporter

MIME_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


class XLSExporter(Exporter):

    def make_download(self, es_dump_path):
        # Set up workbook
        self.workbook = Workbook(write_only=True)
        create_sheet = self.workbook.create_sheet
        for entity_type in ('location', 'party', 'tenure_rel'):
            metadatum = self.metadata[entity_type]
            worksheet = create_sheet(title=metadatum['title'])
            worksheet.append(list(metadatum['attr_columns']))
            metadatum['worksheet'] = worksheet

        # Process ES dump file
        f = open(es_dump_path, encoding='utf-8')
        while True:
            # Read 2 lines in the dump file
            type_line = f.readline()
            source_line = f.readline()
            if not type_line:
                break
            self.process_entity(type_line, source_line, self.write_xls_row)

        # Finalize
        xls_path = os.path.splitext(es_dump_path)[0] + '.xlsx'
        self.workbook.save(filename=xls_path)
        return xls_path, MIME_TYPE

    def write_xls_row(self, entity, metadatum):
        values = self.get_attr_values(entity, metadatum)
        data = metadatum['attr_columns'].copy()
        data.update(values)
        metadatum['worksheet'].append(list(data.values()))
